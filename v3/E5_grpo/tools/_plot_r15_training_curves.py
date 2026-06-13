"""R15 DAPO training curves (9-panel combined figure).

Reads R15_baseit_r15_verl_dapo_full_15ep_summary.xlsx and plots:
  Row 1 (training stability):   loss / grad_norm / lr
  Row 2 (RL dynamics):          judge_mean / entropy / clip_ratio + pg_clipfrac
  Row 3 (capability + format):  dev/test pass@1 / boxed_rate / resp_len + reward dist
"""
import openpyxl
from pathlib import Path

import matplotlib.font_manager as _fm
import matplotlib.pyplot as plt
import os
_msyh_paths = ["/mnt/c/Windows/Fonts/msyh.ttc", "C:\\Windows\\Fonts\\msyh.ttc", "C:/Windows/Fonts/msyh.ttc"]
for _p in _msyh_paths:
    if os.path.exists(_p):
        _fm.fontManager.addfont(_p); break
import numpy as np

ROOT = Path("/mnt/d/fine-tuning")
XLSX = ROOT / "v3/E5_grpo/outputs/R15_baseit_r15_verl_dapo_full_15ep_summary.xlsx"
OUT = ROOT / "v3/E5_grpo/outputs/R15_training_curves.png"

plt.rcParams.update({
    "font.sans-serif": ["Microsoft YaHei", "DejaVu Sans"],
    "axes.unicode_minus": False,
    "font.size": 10,
    "axes.spines.top": False, "axes.spines.right": False,
})


def load_sheet(ws):
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {headers[i]: ws.cell(r, i + 1).value for i in range(len(headers))}
        if row.get("step") is not None:
            rows.append(row)
    return rows


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    train = load_sheet(wb["train_metrics"])
    dev = load_sheet(wb["dev_eval"])
    test = load_sheet(wb["test_eval"])
    reward = load_sheet(wb["reward_dist"])
    summary = {wb["summary"].cell(r, 1).value: wb["summary"].cell(r, 2).value
               for r in range(2, wb["summary"].max_row + 1)
               if wb["summary"].cell(r, 1).value}

    # Helpers
    def col(rows, key):
        return [r[key] for r in rows if r.get(key) is not None]
    def steps(rows):
        return [r["step"] for r in rows]

    base_dev = float(summary.get("base dev pass@1", 70.4))
    base_test = float(summary.get("base test pass@1", 61.94))

    # ===== Figure =====
    fig, axes = plt.subplots(3, 3, figsize=(16, 12))

    # === Row 1: Training stability ===
    # (1,1) loss
    ax = axes[0, 0]
    ax.plot(steps(train), col(train, "loss"), "o-", color="#2563eb", linewidth=1.5, markersize=4)
    ax.set_xlabel("step")
    ax.set_ylabel("policy gradient loss")
    ax.set_title("L1.1 — Train Loss\n小幅波动属正常, GRPO/DAPO 的 advantage 是 normalized 形式",
                 loc="left", fontsize=10, fontweight="semibold")
    ax.grid(alpha=0.3, linestyle=":")
    ax.axhline(0, color="black", linewidth=0.5, alpha=0.5)

    # (1,2) grad_norm
    ax = axes[0, 1]
    ax.plot(steps(train), col(train, "grad_norm"), "o-", color="#dc2626", linewidth=1.5, markersize=4)
    ax.set_xlabel("step"); ax.set_ylabel("grad norm")
    ax.set_title("L1.2 — Gradient Norm\n稳定 → 训练健康, spike → 不稳",
                 loc="left", fontsize=10, fontweight="semibold")
    ax.grid(alpha=0.3, linestyle=":")

    # (1,3) lr schedule
    ax = axes[0, 2]
    ax.plot(steps(train), col(train, "lr"), "o-", color="#16a34a", linewidth=1.5, markersize=4)
    ax.set_xlabel("step"); ax.set_ylabel("learning rate")
    ax.set_title(f"L1.3 — Learning Rate Schedule\nwarmup=7 → 2e-5 const",
                 loc="left", fontsize=10, fontweight="semibold")
    ax.grid(alpha=0.3, linestyle=":")
    ax.ticklabel_format(style="sci", axis="y", scilimits=(0, 0))

    # === Row 2: RL dynamics ===
    # (2,1) judge_mean (reward signal)
    ax = axes[1, 0]
    ax.plot(steps(train), col(train, "judge_mean"), "o-", color="#7c3aed", linewidth=1.5, markersize=4)
    ax.set_xlabel("step"); ax.set_ylabel("judge mean reward")
    ax.set_title(f"L2.1 — Reward (rule judge, in [0,1])\n上升 → policy 学到 correct format + 正确率",
                 loc="left", fontsize=10, fontweight="semibold")
    ax.grid(alpha=0.3, linestyle=":")
    ax.axhline(0.5, color="gray", linewidth=0.5, linestyle="--", alpha=0.5)

    # (2,2) entropy + clip_ratio
    ax = axes[1, 1]
    ax2 = ax.twinx()
    ax.plot(steps(train), col(train, "entropy"), "o-", color="#0891b2", linewidth=1.5, markersize=4, label="entropy")
    ax2.plot(steps(train), col(train, "clip_ratio"), "s-", color="#ea580c", linewidth=1.5, markersize=4, label="clip_ratio")
    ax.set_xlabel("step")
    ax.set_ylabel("policy entropy", color="#0891b2")
    ax2.set_ylabel("PPO clip ratio (importance sampling clip 占比)", color="#ea580c")
    ax.tick_params(axis="y", labelcolor="#0891b2")
    ax2.tick_params(axis="y", labelcolor="#ea580c")
    ax.set_title(f"L2.2 — Entropy ↓ vs Clip ratio ↑\nentropy ↓ = sharpening · clip ↑ = IS drift 拉大",
                 loc="left", fontsize=10, fontweight="semibold")
    ax.grid(alpha=0.3, linestyle=":")

    # (2,3) pg_clipfrac + num_gen_batch
    ax = axes[1, 2]
    ax2 = ax.twinx()
    ax.plot(steps(train), col(train, "pg_clipfrac"), "o-", color="#be185d", linewidth=1.5, markersize=4)
    ax2.plot(steps(train), col(train, "num_gen_batch"), "s-", color="#0891b2", linewidth=1.5, markersize=4)
    ax.set_xlabel("step")
    ax.set_ylabel("pg_clipfrac (clip 占样本比例)", color="#be185d")
    ax2.set_ylabel("filter_groups oversample (×)", color="#0891b2")
    ax.tick_params(axis="y", labelcolor="#be185d")
    ax2.tick_params(axis="y", labelcolor="#0891b2")
    ax.set_title(f"L2.3 — DAPO Clip-Higher 触发率 + DS Oversample\nclipfrac 涨 → ε_high 起作用",
                 loc="left", fontsize=10, fontweight="semibold")
    ax.grid(alpha=0.3, linestyle=":")

    # === Row 3: Capability + format ===
    # (3,1) Dev + Test pass@1 over steps
    ax = axes[2, 0]
    ax.plot(steps(dev), col(dev, "pass@1 (%)"), "o-", color="#2563eb", linewidth=2,
            markersize=5, label=f"dev (base={base_dev:.1f}%)")
    ax.plot(steps(test), col(test, "pass@1 (%)"), "s-", color="#dc2626", linewidth=2,
            markersize=5, label=f"test (base={base_test:.1f}%)")
    ax.axhline(base_dev, color="#2563eb", linestyle="--", alpha=0.4, linewidth=1)
    ax.axhline(base_test, color="#dc2626", linestyle="--", alpha=0.4, linewidth=1)
    # mark best ck-15
    best_step = 15
    dev_best = [r["pass@1 (%)"] for r in dev if r["step"] == best_step][0]
    test_best = [r["pass@1 (%)"] for r in test if r["step"] == best_step][0]
    ax.scatter([best_step], [dev_best], s=200, marker="*", color="#fbbf24",
               edgecolor="black", zorder=10, label=f"ck-{best_step} (dev best)")
    ax.set_xlabel("step (ckpt)"); ax.set_ylabel("pass@1 (%)")
    ax.set_title(f"L3.1 — GSM8K Pass@1 Trajectory\n虚线 = base IT 起点, ★ = dev-select best",
                 loc="left", fontsize=10, fontweight="semibold")
    ax.legend(fontsize=9, loc="lower right")
    ax.grid(alpha=0.3, linestyle=":")

    # (3,2) Boxed rate
    ax = axes[2, 1]
    ax.plot(steps(dev), col(dev, "boxed (%)"), "o-", color="#2563eb", linewidth=2, markersize=5, label="dev")
    ax.plot(steps(test), col(test, "boxed (%)"), "s-", color="#dc2626", linewidth=2, markersize=5, label="test")
    ax.set_xlabel("step (ckpt)"); ax.set_ylabel("boxed_accuracy (%)")
    ax.set_title(f"L3.2 — Boxed Accuracy (strict boxed{{}} 答对)\nbase ~30%, DAPO 训练后涨到 ~70%+",
                 loc="left", fontsize=10, fontweight="semibold")
    ax.legend(fontsize=9, loc="lower right")
    ax.grid(alpha=0.3, linestyle=":")

    # (3,3) Response length + reward distribution
    ax = axes[2, 2]
    ax.plot(steps(dev), col(dev, "mean_len"), "o-", color="#2563eb", linewidth=2, markersize=5, label="dev")
    ax.plot(steps(test), col(test, "mean_len"), "s-", color="#dc2626", linewidth=2, markersize=5, label="test")
    ax.plot(steps(train), col(train, "resp_len"), "x-", color="gray", linewidth=1, markersize=4,
            alpha=0.6, label="train (rollout)")
    ax.set_xlabel("step (ckpt)"); ax.set_ylabel("mean response length (tokens)")
    ax.set_title(f"L3.3 — Output Length (sharpening signature)\n训练中 resp_len ↓ → DAPO 学到更简洁",
                 loc="left", fontsize=10, fontweight="semibold")
    ax.legend(fontsize=9, loc="upper right")
    ax.grid(alpha=0.3, linestyle=":")

    fig.suptitle(
        f"R15 DAPO Training Curves — Gemma2-2B-IT, GSM8K+MATH, 2 epoch, 23 ckpt, 7h 29min\n"
        f"  ck-15 dev-best: dev={dev_best}% (Δ+{dev_best-base_dev:.1f}pp) · test={test_best}% (Δ+{test_best-base_test:.1f}pp vs base IT)",
        fontsize=12, fontweight="semibold", y=1.00,
    )
    plt.tight_layout()
    plt.savefig(OUT, dpi=200, bbox_inches="tight", facecolor="white")
    print(f"saved: {OUT}")

    # Print 关键 numbers
    print("\n=== Key training metrics summary ===")
    print(f"  base dev pass@1:  {base_dev}%")
    print(f"  base test pass@1: {base_test}%")
    print(f"  ck-15 dev:        {dev_best}%  (Δ+{dev_best-base_dev:.1f}pp)")
    print(f"  ck-15 test:       {test_best}%  (Δ+{test_best-base_test:.1f}pp)")
    print(f"  ck-15 train metrics:")
    for r in train:
        if r["step"] == 15 * 10 or r["step"] >= 140:  # ~ck-15 is around step 140-160
            print(f"    step={r['step']}: loss={r['loss']:.4f}, entropy={r['entropy']:.3f}, "
                  f"judge={r['judge_mean']:.3f}, resp_len={r['resp_len']:.0f}, "
                  f"clipfrac={r.get('pg_clipfrac', 0):.4f}")
            break
    print(f"\n  Total train steps logged: {len(train)}")
    print(f"  Total dev eval ckpt: {len(dev)}")
    print(f"  Total test eval ckpt: {len(test)}")


if __name__ == "__main__":
    main()
